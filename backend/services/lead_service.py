from __future__ import annotations

import csv
import io
import re
from datetime import datetime, date, timezone
from typing import List, Optional, Tuple

from fastapi import HTTPException
from sqlalchemy import or_, and_, func, cast, Date, String, case
from sqlalchemy.orm import Session, joinedload

from models.lead import Lead, LeadStatus, LeadType
from models.lead_remark import LeadRemark
from models.user import User, UserRole
from schemas.lead_schema import LeadCreate, LeadUpdate, PipelineMove
from services import activity_service


# --- Pipeline: one step at a time; LOST from any open stage; CONVERTED only from FOLLOW_UP ---

PIPELINE_ORDER: List[LeadStatus] = [
    LeadStatus.NEW,
    LeadStatus.CONTACTED,
    LeadStatus.INTERESTED,
    LeadStatus.FOLLOW_UP,
    LeadStatus.CONVERTED,
]
TERMINAL = frozenset({LeadStatus.CONVERTED, LeadStatus.LOST})


def _pipeline_index(s: LeadStatus) -> Optional[int]:
    try:
        return PIPELINE_ORDER.index(s)
    except ValueError:
        return None


def validate_pipeline_transition(old: LeadStatus, new: LeadStatus) -> None:
    if old == new:
        return
    # UI requirement: pipeline drag-drop can move a lead to any stage.
    return


def can_access_lead(user: User, lead: Lead) -> bool:
    if user.role in (UserRole.ADMIN, UserRole.MANAGER):
        return True
    return lead.assigned_to == user.id


def get_lead(db: Session, lead_id: int) -> Optional[Lead]:
    return db.query(Lead).filter(Lead.id == lead_id, Lead.is_deleted == False).first()


def get_lead_for_user(db: Session, lead_id: int, user: User) -> Lead:
    lead = get_lead(db, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")
    if not can_access_lead(user, lead):
        raise HTTPException(status_code=403, detail="Not allowed to access this lead")
    return lead


def _scalar_for_compare(v):
    if v is None:
        return None
    if hasattr(v, "value"):
        return v.value
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if isinstance(v, float):
        return round(v, 6)
    return v


def _activity_val(x):
    """Human-readable value for activity log lines (enum → string value)."""
    if x is None:
        return None
    if hasattr(x, "value"):
        return x.value
    if hasattr(x, "isoformat"):
        return x.isoformat()
    return x


def _field_changed(old, new) -> bool:
    return _scalar_for_compare(old) != _scalar_for_compare(new)


def _payment_ok(amount: float | None, method: str | None) -> bool:
    if amount is None or (isinstance(amount, (int, float)) and float(amount) <= 0):
        return False
    if not method or not str(method).strip():
        return False
    return True


def _naive_utc(dt: Optional[datetime]) -> Optional[datetime]:
    """TIMESTAMP WITHOUT TIME ZONE columns: psycopg rejects tz-aware values unless normalized."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def _lead_status_text():
    """Cast for SQL: PG native leadstatus enum vs VARCHAR bind params from TypeDecorator."""
    return func.upper(cast(Lead.status, String))


def _lead_type_text():
    """Cast for SQL: PG native leadtype enum vs VARCHAR bind params from TypeDecorator."""
    return func.upper(cast(Lead.lead_type, String))


def _terminal_status_strings() -> list[str]:
    return ["CONVERTED", "NOT_INTERESTED"]


def search_leads(
    db: Session,
    user: User,
    *,
    skip: int = 0,
    limit: int = 50,
    q: Optional[str] = None,
    status: Optional[LeadStatus] = None,
    assigned_to: Optional[int] = None,
    lead_type: Optional[LeadType] = None,
    source: Optional[str] = None,
    created_from: Optional[datetime] = None,
    created_to: Optional[datetime] = None,
       overdue_only: bool = False,
    follow_up_on: Optional[date] = None,
) -> Tuple[List[Lead], int]:
    query = db.query(Lead).filter(Lead.is_deleted == False)
    if user.role == UserRole.SALES_AGENT:
        query = query.filter(Lead.assigned_to == user.id)

    if q and q.strip():
        term = f"%{q.strip()}%"
        query = query.filter(
            or_(
                Lead.name.ilike(term),
                Lead.email.ilike(term),
                Lead.phone.ilike(term),
            )
        )
    if status is not None:
        target = "NOT_INTERESTED" if status == LeadStatus.LOST else status.value.replace("-", "_").upper()
        query = query.filter(_lead_status_text() == target)
    if assigned_to is not None:
        query = query.filter(Lead.assigned_to == assigned_to)
    if lead_type is not None:
        query = query.filter(_lead_type_text() == lead_type.value.upper())
    if source is not None and str(source).strip():
        term = f"%{str(source).strip()}%"
        query = query.filter(Lead.source.isnot(None), Lead.source.ilike(term))
    if created_from is not None:
        query = query.filter(Lead.created_at >= created_from)
    if created_to is not None:
        query = query.filter(Lead.created_at <= created_to)

    now = datetime.utcnow()
    time_clauses = []
    if overdue_only:
        time_clauses.append(
            and_(
                Lead.follow_up_date.isnot(None),
                Lead.follow_up_date < now,
                _lead_status_text().notin_(_terminal_status_strings()),
            )
        )
    if follow_up_on is not None:
        time_clauses.append(
            and_(
                Lead.follow_up_date.isnot(None),
                cast(Lead.follow_up_date, Date) == follow_up_on,
            )
        )
    if len(time_clauses) >= 2:
        query = query.filter(or_(*time_clauses))
    elif len(time_clauses) == 1:
        query = query.filter(time_clauses[0])

    total = query.count()
    prio = case(
        (Lead.priority == "HOT", 0),
        (Lead.priority == "WARM", 1),
        (Lead.priority == "COLD", 2),
        else_=3,
    )
    rows = (
        query.order_by(prio.asc(), Lead.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return rows, total


def list_lead_remarks(db: Session, lead_id: int):
    return (
        db.query(LeadRemark)
        .options(joinedload(LeadRemark.user))
        .filter(LeadRemark.lead_id == lead_id)
        .order_by(LeadRemark.created_at.desc())
        .all()
    )


def add_lead_remark(db: Session, lead_id: int, user_id: int, data: lead_schema.LeadRemarkCreate) -> LeadRemark:
    text = (data.body or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Remark cannot be empty")
    
    # Use provided timestamp or let DB default handle it
    created_at = data.created_at
    if created_at and created_at.tzinfo is not None:
         # Normalize to UTC naive if the column isn't aware yet, 
         # but I changed it to aware. So we just ensure it's UTC.
         created_at = created_at.astimezone(timezone.utc)

    r = LeadRemark(lead_id=lead_id, user_id=user_id, body=text)
    if created_at:
        r.created_at = created_at

    db.add(r)
    db.commit()
    activity_service.log_activity(
        db,
        lead_id=lead_id,
        user_id=user_id,
        action="Added remark",
        details=text[:2000],
    )
    return (
        db.query(LeadRemark)
        .options(joinedload(LeadRemark.user))
        .filter(LeadRemark.id == r.id)
        .one()
    )


def get_leads(db: Session, skip: int = 0, limit: int = 100, user: Optional[User] = None):
    """Used for CSV export and legacy callers; respects sales scope when user passed."""
    q = db.query(Lead).filter(Lead.is_deleted == False)
    if user is not None and user.role == UserRole.SALES_AGENT:
        q = q.filter(Lead.assigned_to == user.id)
    return q.order_by(Lead.created_at.desc()).offset(skip).limit(limit).all()


def get_overdue_leads(db: Session, user: User, skip: int = 0, limit: int = 200):
    now = datetime.utcnow()
    q = (
        db.query(Lead)
        .filter(
            Lead.is_deleted == False,
            Lead.follow_up_date.isnot(None),
            Lead.follow_up_date < now,
            _lead_status_text().notin_(_terminal_status_strings()),
        )
        .order_by(Lead.follow_up_date.asc())
    )
    if user.role == UserRole.SALES_AGENT:
        q = q.filter(Lead.assigned_to == user.id)
    return q.offset(skip).limit(limit).all()


def stats_for_user(db: Session, user: User) -> dict:
    def scope(query):
        if user.role == UserRole.SALES_AGENT:
            return query.filter(Lead.assigned_to == user.id)
        return query

    total = scope(db.query(Lead).filter(Lead.is_deleted == False)).count()
    by_status = (
        scope(db.query(Lead).filter(Lead.is_deleted == False))
        .with_entities(Lead.status, func.count(Lead.id))
        .group_by(Lead.status)
        .all()
    )
    status_summary = {s.value: 0 for s in LeadStatus}
    for st, cnt in by_status:
        status_summary[st.value] = cnt
    converted = status_summary.get(LeadStatus.CONVERTED.value, 0)

    today = date.today()
    base = db.query(Lead).filter(
        Lead.is_deleted == False,
        Lead.follow_up_date.isnot(None),
        cast(Lead.follow_up_date, Date) == today,
    )
    if user.role == UserRole.SALES_AGENT:
        base = base.filter(Lead.assigned_to == user.id)
    followups_today = base.count()

    overdue_q = db.query(Lead).filter(
        Lead.is_deleted == False,
        Lead.follow_up_date.isnot(None),
        Lead.follow_up_date < datetime.utcnow(),
        _lead_status_text().notin_(_terminal_status_strings()),
    )
    if user.role == UserRole.SALES_AGENT:
        overdue_q = overdue_q.filter(Lead.assigned_to == user.id)
    overdue_count = overdue_q.count()

    return {
        "total_leads": total,
        "status_summary": status_summary,
        "converted": converted,
        "followups_today": followups_today,
        "overdue": overdue_count,
    }


def create_lead(db: Session, lead: LeadCreate, user_id: int):
    if lead.status == LeadStatus.INTERESTED and not lead.follow_up_date:
        raise HTTPException(
            status_code=400,
            detail="Follow-up date is required for 'Interested' leads",
        )

    if lead.status == LeadStatus.CONVERTED:
        if not _payment_ok(lead.payment_amount, lead.payment_method):
            raise HTTPException(
                status_code=400,
                detail="Payment amount and payment method are required for 'Converted' leads",
            )

    data = lead.model_dump()
    assignee_id = data.pop("assigned_to", None)
    if assignee_id is None:
        assignee_id = user_id
    target = db.query(User).filter(User.id == assignee_id).first()
    if not target:
        raise HTTPException(status_code=400, detail="Assigned user does not exist")
    data["assigned_to"] = assignee_id

    if lead.status == LeadStatus.CONVERTED:
        data["converted_at"] = datetime.utcnow()

    if data.get("follow_up_date") is not None:
        data["follow_up_date"] = _naive_utc(data["follow_up_date"])

    db_lead = Lead(**data)
    db.add(db_lead)
    db.commit()
    db.refresh(db_lead)

    activity_service.log_activity(
        db,
        lead_id=db_lead.id,
        user_id=user_id,
        action="Lead Created",
        details=f"Lead {db_lead.name} added to the system.",
    )

    return db_lead


def update_lead(db: Session, lead_id: int, lead_update: LeadUpdate, user_id: int):
    db_lead = get_lead(db, lead_id)
    if not db_lead:
        return None

    # exclude_unset=True ensures we only update fields provided in the PATCH request
    update_data = lead_update.model_dump(exclude_unset=True)

    # Validate assignee if provided
    new_assignee_id = update_data.get("assigned_to")
    if new_assignee_id is not None:
        target_user = db.query(User).filter(User.id == new_assignee_id).first()
        if not target_user:
            raise HTTPException(
                status_code=400, 
                detail=f"Assigned user with id {new_assignee_id} does not exist"
            )

    new_status = update_data.get("status")
    if new_status is not None:
        try:
            # Pydantic may have already coerced it to enum value if valid
            if isinstance(new_status, LeadStatus):
                coerced_status = new_status
            else:
                s = str(new_status).strip().lower().replace("_", "-")
                coerced_status = LeadStatus(s)
            
            validate_pipeline_transition(db_lead.status, coerced_status)

            if coerced_status == LeadStatus.INTERESTED:
                fu = update_data.get("follow_up_date")
                if fu is None and db_lead.follow_up_date is None:
                    raise HTTPException(
                        status_code=400,
                        detail="Follow-up date is required for 'Interested' leads",
                    )
            if coerced_status == LeadStatus.CONVERTED:
                amt = update_data.get("payment_amount", db_lead.payment_amount)
                meth = update_data.get("payment_method", db_lead.payment_method)
                if not _payment_ok(amt, meth):
                    raise HTTPException(
                        status_code=400,
                        detail="Payment amount and payment method are required for 'Converted' leads",
                    )
                db_lead.converted_at = datetime.utcnow()
        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status value: {new_status}. Error: {str(e)}"
            )

    changes = []
    for key, value in update_data.items():
        if not hasattr(db_lead, key):
            continue
            
        old_val = getattr(db_lead, key)
        if not _field_changed(old_val, value):
            continue
            
        coerced = value
        if key == "follow_up_date" and coerced is not None:
            coerced = _naive_utc(coerced)
        elif key == "status" and value is not None:
            if not isinstance(value, LeadStatus):
                try:
                    s = str(value).strip().lower().replace("_", "-")
                    coerced = LeadStatus(s)
                except ValueError:
                    coerced = value
        elif key == "lead_type" and value is not None:
            if not isinstance(value, LeadType):
                try:
                    s = str(value).strip().lower()
                    coerced = LeadType(s)
                except ValueError:
                    coerced = value
        
        # Additional safety for mandatory fields like 'name'
        if key == "name" and (value is None or str(value).strip() == ""):
             raise HTTPException(status_code=400, detail="Name cannot be empty")

        setattr(db_lead, key, coerced)

        if key == "assigned_to":
            old_n = db.query(User.full_name).filter(User.id == old_val).scalar() if old_val else "Unassigned"
            new_n = db.query(User.full_name).filter(User.id == coerced).scalar() if coerced else "Unassigned"
            changes.append(f"assigned_to: {old_n} → {new_n}")
        else:
            changes.append(f"{key}: {_activity_val(old_val)} → {_activity_val(coerced)}")

    if changes:
        try:
            db.commit()
            db.refresh(db_lead)
            activity_service.log_activity(
                db,
                lead_id=db_lead.id,
                user_id=user_id,
                action="Lead Updated",
                details=" | ".join(changes[:20]),
            )
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail=f"Database update failed: {str(e)}"
            )

    return db_lead


def delete_lead(db: Session, lead_id: int) -> bool:
    db_lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not db_lead:
        return False
    db_lead.is_deleted = True
    db.commit()
    return True


def move_lead_pipeline(
    db: Session, lead_id: int, move: PipelineMove, user_id: int
):
    db_lead = get_lead(db, lead_id)
    if not db_lead:
        return None

    new_status = move.status
    validate_pipeline_transition(db_lead.status, new_status)

    if new_status == LeadStatus.INTERESTED:
        fu = move.follow_up_date or db_lead.follow_up_date
        if fu is None:
            raise HTTPException(
                status_code=400,
                detail="Follow-up date is required when moving to 'Interested'",
            )

    if new_status == LeadStatus.CONVERTED:
        amt = (
            move.payment_amount
            if move.payment_amount is not None
            else db_lead.payment_amount
        )
        meth = (
            move.payment_method
            if move.payment_method is not None
            else db_lead.payment_method
        )
        if not _payment_ok(amt, meth):
            raise HTTPException(
                status_code=400,
                detail="Payment amount and payment method are required for 'Converted'",
            )
        db_lead.payment_amount = float(amt) if amt is not None else db_lead.payment_amount
        db_lead.payment_method = meth
        db_lead.converted_at = datetime.utcnow()

    old_status = db_lead.status
    db_lead.status = new_status
    if move.follow_up_date is not None:
        db_lead.follow_up_date = _naive_utc(move.follow_up_date)

    db.commit()
    db.refresh(db_lead)

    activity_service.log_activity(
        db,
        lead_id=db_lead.id,
        user_id=user_id,
        action="Pipeline Move",
        details=f"Status changed from {_activity_val(old_status)} → {_activity_val(new_status)}",
    )
    return db_lead


def _parse_import_status(raw) -> LeadStatus:
    if raw is None or not str(raw).strip():
        return LeadStatus.NEW
    s = str(raw).strip().lower().replace(" ", "-")
    if s == "follow_up":
        s = "follow-up"
    if s in ("not_interested", "not-interested"):
        s = "lost"
    try:
        return LeadStatus(s)
    except ValueError:
        return LeadStatus.NEW

def is_valid_email(email: str) -> bool:
    return bool(re.match(r"[^@]+@[^@]+\.[^@]+", email)) if email else False

def is_valid_phone(phone: str) -> bool:
    return len(re.sub(r"\D", "", phone)) >= 5 if phone else False

def get_existing_lead(db: Session, email: str | None, phone: str | None) -> Optional[Lead]:
    if not email and not phone:
        return None
    clauses = []
    if email:
        clauses.append(Lead.email == email)
    if phone:
        clauses.append(Lead.phone == phone)
    return db.query(Lead).filter(or_(*clauses)).first()

def import_leads_from_rows(
    db: Session,
    rows: List[dict],
    user_id: int,
    mode: str = "skip",
) -> dict:
    """rows: list of normalized dicts. mode: 'skip' or 'update' existing leads."""
    success = 0
    updated = 0
    failed = 0
    errors: List[str] = []
    
    for i, row in enumerate(rows, start=2):
        # Trim and normalize data
        data = {k: (str(v).strip() if v is not None else None) for k, v in row.items()}
        name = data.get("name")
        phone = data.get("phone")
        email = data.get("email")

        if not name:
            errors.append(f"Row {i}: Name is required")
            failed += 1
            continue
        
        if email and not is_valid_email(email):
            errors.append(f"Row {i}: Invalid email format ({email})")
            failed += 1
            continue

        if phone and not is_valid_phone(phone):
            errors.append(f"Row {i}: Invalid phone number ({phone})")
            failed += 1
            continue

        try:
            # Check for existing
            existing = get_existing_lead(db, email, phone)
            
            if existing:
                if mode == "skip":
                    errors.append(f"Row {i}: Skipped (Duplicate found for {email or phone})")
                    failed += 1
                    continue
                else:
                    # Update mode
                    for k, v in data.items():
                        if hasattr(existing, k) and v is not None:
                            setattr(existing, k, v)
                    db.commit()
                    updated += 1
                    activity_service.log_activity(
                        db,
                        lead_id=existing.id,
                        user_id=user_id,
                        action="Lead Updated",
                        details=f"Import update: {existing.name}",
                    )
                    continue

            # Create new
            st = _parse_import_status(data.get("status"))
            lead = Lead(
                name=name,
                phone=phone,
                email=email or None,
                company_name=data.get("company_name"),
                website_url=data.get("website_url"),
                linkedin_url=data.get("linkedin_url"),
                location=data.get("location"),
                category=data.get("category"),
                industry=data.get("industry"),
                notes=data.get("notes"),
                status=st,
                assigned_to=user_id, # Default to importer
                lead_type=LeadType.INBOUND,
            )
            db.add(lead)
            db.commit()
            db.refresh(lead)
            success += 1
            activity_service.log_activity(
                db,
                lead_id=lead.id,
                user_id=user_id,
                action="Lead Created",
                details=f"Imported: {lead.name}",
            )
        except Exception as e:
            db.rollback()
            errors.append(f"Row {i}: {e!s}")
            failed += 1

    return {
        "total": len(rows),
        "success": success,
        "updated": updated,
        "failed": failed,
        "errors": errors[:50], # Limit error log size
    }


def normalize_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_")


def parse_import_file(content: bytes, filename: str) -> List[dict]:
    name = (filename or "").lower()
    rows = []
    
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            m = {normalize_header(k): (v or "").strip() for k, v in r.items() if k}
            if not any(m.values()): continue # skip empty
            rows.append(m)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel support requires openpyxl")
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [normalize_header(str(c.value) if c.value else "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            m = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers)) if i < len(headers)}
            rows.append(m)
    else:
        raise HTTPException(status_code=400, detail="Use .csv or .xlsx file")

    final = []
    for m in rows:
        mapped = {
            "name": m.get("name") or m.get("full_name") or m.get("lead_name") or m.get("client_name") or m.get("customer"),
            "phone": m.get("phone") or m.get("mobile") or m.get("contact") or m.get("tel") or m.get("whatsapp"),
            "email": m.get("email") or m.get("e_mail") or m.get("mail"),
            "company_name": m.get("company_name") or m.get("company") or m.get("organization") or m.get("firm"),
            "website_url": m.get("website_url") or m.get("website") or m.get("url") or m.get("site"),
            "linkedin_url": m.get("linkedin_url") or m.get("linkedin") or m.get("linked_in"),
            "location": m.get("location") or m.get("city") or m.get("address") or m.get("country") or m.get("state"),
            "category": m.get("category") or m.get("segment") or m.get("group") or m.get("tag"),
            "industry": m.get("industry") or m.get("sector") or m.get("vertical") or m.get("business"),
            "status": m.get("status"),
            "notes": m.get("notes") or m.get("remarks") or m.get("comments") or m.get("description"),
        }
        final.append(mapped)
    
    return final


SAMPLE_CSV = """name,phone,email,company_name,status,assigned_to,source,location,notes,budget
Acme Demo,+15551234567,demo@acme.com,Acme Inc,new,,website,NY,
"""
